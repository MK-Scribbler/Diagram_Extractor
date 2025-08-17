import os
import cv2
import shutil
import requests
from flask import Flask, request, render_template, redirect, url_for
from urllib.parse import urljoin, urlparse
from PIL import Image
import io
from ultralytics import YOLO
import time
import base64
from playwright.sync_api import sync_playwright
from werkzeug.exceptions import RequestEntityTooLarge
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
import numpy as np

# Flask app initialization and configuration
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 1024 * 1024 * 1024  # 1GB max upload size

# === Configurable paths and model settings ===
UPLOAD_FOLDER = 'uploads'
WEB_OUTPUT_FOLDER = 'static/diagrams_web_filtered'
PDF_OUTPUT_FOLDER = 'static/diagrams_pdf_filtered'
MODEL_PATH = "best.pt"
CONFIDENCE_THRESHOLD = 0.6
IOU_THRESHOLD = 0.8
PADDING = 25

# --- Utility: Log image details to Excel ---
def log_image_to_excel(image_name, source, excel_path='image_log.xlsx'):
    """Log the image name and its source to an Excel file for tracking."""
    results_dir = os.path.join('results', 'Diagrams_Extracted')
    image_path = os.path.join(results_dir, image_name)
    if not os.path.exists(image_path):
        print(f"Image {image_path} not found, skipping log.")
        return
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(['Image Name', 'Source'])
        wb.save(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active
    existing_names = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        existing_names.add(row[0])
    if image_name in existing_names:
        print(f"Duplicate image {image_name}, skipping log.")
        return
    ws.append([image_name, source])
    wb.save(excel_path)
    print(f"Logged {image_name}, source: {source}.")

# --- Ensure output folders exist ---
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(WEB_OUTPUT_FOLDER, exist_ok=True)
os.makedirs(PDF_OUTPUT_FOLDER, exist_ok=True)

# --- Load YOLO model for diagram detection ---
model = YOLO(MODEL_PATH)

# --- Jinja2 filter for extracting basename from file path ---
@app.template_filter('basename')
def basename_filter(value):
    return os.path.basename(value)

# --- Utility: Calculate Intersection over Union (IoU) for bounding boxes ---
def iou(boxA, boxB):
    xa1, ya1, xa2, ya2 = boxA
    xb1, yb1, xb2, yb2 = boxB
    inter_x1 = max(xa1, xb1)
    inter_y1 = max(ya1, yb1)
    inter_x2 = min(xa2, xb2)
    inter_y2 = min(ya2, yb2)
    inter_area = max(0, inter_x2 - inter_x1) * max(0, inter_y2 - inter_y1)
    areaA = (xa2 - xa1) * (ya2 - ya1)
    areaB = (xb2 - xb1) * (yb2 - yb1)
    union_area = areaA + areaB - inter_area
    return inter_area / union_area if union_area > 0 else 0

# --- Home page route ---
@app.route('/')
def index():
    return render_template('index.html')

# --- Scrape images from a web page using Playwright ---
@app.route('/', methods=['POST'])
def scrape_images():
    url = request.form.get('url')
    if not url:
        return redirect(url_for('index'))
    # Ensure URL has protocol
    if not url.startswith('http://') and not url.startswith('https://'):
        url = 'https://' + url

    shutil.rmtree(WEB_OUTPUT_FOLDER, ignore_errors=True)
    os.makedirs(WEB_OUTPUT_FOLDER, exist_ok=True)

    sources = []
    logo_keywords = ['logo', 'icon', 'sprite', 'favicon', 'avatar', 'profile', 'thumb', 'placeholder']

    def is_google_images(url):
        return "tbm=isch" in url or "google.com/search" in url

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_default_navigation_timeout(0)
        page.set_default_timeout(60_000)

        try:
            page.goto(url, wait_until="domcontentloaded")
        except Exception as nav_err:
            print(f"[WARN] first navigation timed-out -> retrying: {nav_err}")
            try:
                page.goto(url, wait_until="commit")
            except Exception as second_err:
                browser.close()
                return render_template(
                    "message.html",
                    status="error",
                    message="⚠️ Could not load that page",
                    redirect_url=url_for('index')
                )

        try:
            page.wait_for_selector("img", state="attached", timeout=20000)
        except Exception:
            print("Warning: No <img> tag found or took too long to load.")
        time.sleep(2)

        # Scroll to load more images
        for _ in range(5):
            page.mouse.wheel(0, 2000)
            page.wait_for_timeout(800)

        count = 1
        if is_google_images(url):
            thumbnails = page.query_selector_all("img.Q4LuWd, img")
            saved_srcs = set()
            for thumb in thumbnails:
                try:
                    thumb.scroll_into_view_if_needed()
                    thumb.click()
                    page.wait_for_timeout(800)
                    images = page.query_selector_all("img.n3VNCb, img")
                    for img in images:
                        src = img.get_attribute("src")
                        if not src:
                            continue
                        src = urljoin(url, src)
                        if not src.lower().startswith("http"):
                            continue
                        if any(word in src.lower() for word in logo_keywords):
                            continue
                        if src in saved_srcs:
                            continue
                        try:
                            img_data = requests.get(src, timeout=5).content
                            img_pil = Image.open(io.BytesIO(img_data))
                            w, h = img_pil.size
                            if w < 80 or h < 80:
                                continue
                            ext = os.path.splitext(urlparse(src).path)[-1] or ".jpg"
                            out_path = os.path.join(WEB_OUTPUT_FOLDER, f"{count}{ext}")
                            with open(out_path, "wb") as f:
                                f.write(img_data)
                            saved_srcs.add(src)
                            sources.append(src)
                            count += 1
                            break
                        except Exception:
                            continue
                except Exception as e:
                    print(f"Error processing thumbnail: {e}")
        else:
            images = page.query_selector_all("img")
            for img in images:
                src = img.get_attribute("src")
                if not src:
                    continue
                src = urljoin(url, src)
                if not src.lower().startswith("http"):
                    continue
                if any(word in src.lower() for word in logo_keywords):
                    continue
                try:
                    img_data = requests.get(src, timeout=5).content
                    img_pil = Image.open(io.BytesIO(img_data))
                    w, h = img_pil.size
                    if w < 80 or h < 80:
                        continue
                    ext = os.path.splitext(urlparse(src).path)[-1] or ".jpg"
                    out_path = os.path.join(WEB_OUTPUT_FOLDER, f"{count}{ext}")
                    with open(out_path, "wb") as f:
                        f.write(img_data)
                    sources.append(src)
                    count += 1
                except Exception:
                    continue
        browser.close()

    saved_files = sorted(os.listdir(WEB_OUTPUT_FOLDER))
    if not saved_files:
        return render_template(
            "message.html",
            status="error",
            message="⚠️ No valid images were found on the page.",
            redirect_url=url_for("index"),
            redirect_message="Try another website"
        )

    image_urls = [url_for('static', filename=f'diagrams_web_filtered/{img}') for img in saved_files]
    return render_template('preview.html', image_urls=image_urls, sources=sources)


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 1024 * 1024 * 1024  # 1GB, adjust as needed


# === Config ===
UPLOAD_FOLDER = 'uploads'
WEB_OUTPUT_FOLDER = 'static/diagrams_web_filtered'
PDF_OUTPUT_FOLDER = 'static/diagrams_pdf_filtered'
POPPLER_PATH = r"C:\\poppler\\poppler-24.08.0\\Library\\bin"
MODEL_PATH = "best.pt"
CONFIDENCE_THRESHOLD = 0.6
IOU_THRESHOLD = 0.8
PADDING = 25


def log_image_to_excel(image_name, source, excel_path='image_log.xlsx'):
    # Corrected path to match the actual save directory
    results_dir = os.path.join('results', 'Diagrams_Extracted')
    image_path = os.path.join(results_dir, image_name)
    if not os.path.exists(image_path):
        print(f"Image {image_path} not found, skipping log.")
        return
    if not os.path.exists(excel_path):
        print("Excel file does not exist, creating...")
        wb = Workbook()
        ws = wb.active
        ws.append(['Image Name', 'Source'])
        wb.save(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active
    existing_names = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        existing_names.add(row[0])
    if image_name in existing_names:
        print(f"Duplicate image {image_name}, skipping log.")
        return
    ws.append([image_name, source])
    wb.save(excel_path)
    print(f"Logged {image_name}, source: {source}.")


os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(WEB_OUTPUT_FOLDER, exist_ok=True)
os.makedirs(PDF_OUTPUT_FOLDER, exist_ok=True)

model = YOLO(MODEL_PATH)


@app.template_filter('basename')
def basename_filter(value):
    return os.path.basename(value)


def iou(boxA, boxB):
    xa1, ya1, xa2, ya2 = boxA
    xb1, yb1, xb2, yb2 = boxB
    inter_x1 = max(xa1, xb1)
    inter_y1 = max(ya1, yb1)
    inter_x2 = min(xa2, xb2)
    inter_y2 = min(ya2, yb2)
    inter_area = max(0, inter_x2 - inter_x1) * max(0, inter_y2 - inter_y1)
    areaA = (xa2 - xa1) * (ya2 - ya1)
    areaB = (xb2 - xb1) * (yb2 - yb1)
    union_area = areaA + areaB - inter_area
    return inter_area / union_area if union_area > 0 else 0


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/', methods=['POST'])
def scrape_images():
    url = request.form.get('url')
    if not url:
        return redirect(url_for('index'))
    # Fix: Ensure URL has protocol
    if not url.startswith('http://') and not url.startswith('https://'):
        url = 'https://' + url

    shutil.rmtree(WEB_OUTPUT_FOLDER, ignore_errors=True)
    os.makedirs(WEB_OUTPUT_FOLDER, exist_ok=True)

    sources = []
    logo_keywords = ['logo', 'icon', 'sprite', 'favicon', 'avatar', 'profile', 'thumb', 'placeholder']

    def is_google_images(url):
        return "tbm=isch" in url or "google.com/search" in url

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_default_navigation_timeout(0)
        page.set_default_timeout(60_000)

        try:
            page.goto(url, wait_until="domcontentloaded")
        except Exception as nav_err:
            print(f"[WARN] first navigation timed-out -> retrying: {nav_err}")
            try:
                page.goto(url, wait_until="commit")
            except Exception as second_err:
                browser.close()
                return render_template(
                    "message.html",
                    status="error",
                    message="⚠️ Could not load that page",
                    redirect_url=url_for('index')
                )

        try:
            page.wait_for_selector("img", state="attached", timeout=20000)
        except Exception:
            print("Warning: No <img> tag found or took too long to load.")
        time.sleep(2)

        for _ in range(5):
            page.mouse.wheel(0, 2000)
            page.wait_for_timeout(800)

        count = 1
        if is_google_images(url):
            thumbnails = page.query_selector_all("img.Q4LuWd, img")
            saved_srcs = set()
            for thumb in thumbnails:
                try:
                    thumb.scroll_into_view_if_needed()
                    thumb.click()
                    page.wait_for_timeout(800)
                    images = page.query_selector_all("img.n3VNCb, img")
                    for img in images:
                        src = img.get_attribute("src")
                        if not src:
                            continue
                        src = urljoin(url, src)
                        if not src.lower().startswith("http"):
                            continue
                        if any(word in src.lower() for word in logo_keywords):
                            continue
                        if src in saved_srcs:
                            continue
                        try:
                            img_data = requests.get(src, timeout=5).content
                            img_pil = Image.open(io.BytesIO(img_data))
                            w, h = img_pil.size
                            if w < 80 or h < 80:
                                continue
                            ext = os.path.splitext(urlparse(src).path)[-1] or ".jpg"
                            out_path = os.path.join(WEB_OUTPUT_FOLDER, f"{count}{ext}")
                            with open(out_path, "wb") as f:
                                f.write(img_data)
                            saved_srcs.add(src)
                            sources.append(src)
                            count += 1
                            break
                        except Exception:
                            continue
                except Exception as e:
                    print(f"Error processing thumbnail: {e}")
        else:
            images = page.query_selector_all("img")
            for img in images:
                src = img.get_attribute("src")
                if not src:
                    continue
                src = urljoin(url, src)
                if not src.lower().startswith("http"):
                    continue
                if any(word in src.lower() for word in logo_keywords):
                    continue
                try:
                    img_data = requests.get(src, timeout=5).content
                    img_pil = Image.open(io.BytesIO(img_data))
                    w, h = img_pil.size
                    if w < 80 or h < 80:
                        continue
                    ext = os.path.splitext(urlparse(src).path)[-1] or ".jpg"
                    out_path = os.path.join(WEB_OUTPUT_FOLDER, f"{count}{ext}")
                    with open(out_path, "wb") as f:
                        f.write(img_data)
                    sources.append(src)
                    count += 1
                except Exception:
                    continue
        browser.close()

    saved_files = sorted(os.listdir(WEB_OUTPUT_FOLDER))
    if not saved_files:
        return render_template(
            "message.html",
            status="error",
            message="⚠️ No valid images were found on the page.",
            redirect_url=url_for("index"),
            redirect_message="Try another website"
        )

    image_urls = [url_for('static', filename=f'diagrams_web_filtered/{img}') for img in saved_files]
    return render_template('preview.html', image_urls=image_urls, sources=sources)



# --- API: Save an image (cropped or original) to results folder and log to Excel ---
@app.route('/save_image', methods=['POST'])
def save_image():
    data = request.json
    relative_path = data.get('image_path')
    source_url = data.get('source')
    cropped_data = data.get('cropped_data')

    if not relative_path:
        return {"status": "error", "message": "Missing image path"}, 400

    extracted_dir = os.path.join('results', 'Diagrams_Extracted')
    os.makedirs(extracted_dir, exist_ok=True)

    # Find the next sequential number for the filename
    existing_numbers = [int(name) for name, _ in (os.path.splitext(f) for f in os.listdir(extracted_dir)) if name.isdigit()]
    next_number = max(existing_numbers, default=19999) + 1

    ext = os.path.splitext(relative_path)[-1]
    dst_filename = f"{next_number}{ext}"
    dst_path = os.path.join(extracted_dir, dst_filename)

    # Save cropped image if available, else copy original
    if cropped_data and cropped_data.startswith("data:image"):
        header, encoded = cropped_data.split(",", 1)
        img_bytes = base64.b64decode(encoded)
        with open(dst_path, "wb") as f:
            f.write(img_bytes)
    else:
        src_path = relative_path.strip('/')
        if os.path.exists(src_path):
            shutil.copyfile(src_path, dst_path)
        else:
            return {"status": "error", "message": "Source file not found"}, 404

    # Log the saved image to Excel
    log_image_to_excel(dst_filename, source_url)
    return {"status": "success", "message": f"Image saved as {dst_filename}", "filename": dst_filename}



# --- API: Delete a saved image and its log entry from Excel ---
@app.route('/delete_saved_image', methods=['POST'])
def delete_saved_image():
    data = request.json
    filename = data.get('filename')

    if not filename:
        return {"status": "error", "message": "Missing filename"}, 400

    extracted_dir = os.path.join('results', 'Diagrams_Extracted')
    file_path = os.path.join(extracted_dir, filename)
    file_deleted = False
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            file_deleted = True
        except Exception as e:
            return {"status": "error", "message": f"Error deleting file: {str(e)}"}, 500

    # Remove entry from Excel log
    excel_path = 'image_log.xlsx'
    log_deleted = False
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            # Find the row to delete by iterating backwards
            for i in range(ws.max_row, 1, -1):
                if ws.cell(row=i, column=1).value == filename:
                    ws.delete_rows(i)
                    wb.save(excel_path)
                    log_deleted = True
                    break
        except Exception as e:
            print(f"Could not delete log for {filename}: {e}")

    if not file_deleted and not log_deleted:
        return {"status": "not_found", "message": "File and log entry not found"}, 404

    return {"status": "success", "message": "File and log entry deleted"}



# --- Error handler for large file uploads ---
@app.errorhandler(RequestEntityTooLarge)
def handle_large_file(e):
    return render_template("message.html", status="error",
                           message="⚠️ The file is too large. Maximum allowed size is 1GB.",
                           redirect_url=url_for('index')), 413



# --- API: Upload a PDF, extract diagrams using YOLO, and save results ---
@app.route('/upload_pdf', methods=['POST'])
def upload_pdf():
    pdf_file = request.files.get('pdf')
    if not pdf_file:
        return redirect(url_for('index'))

    shutil.rmtree(PDF_OUTPUT_FOLDER, ignore_errors=True)
    os.makedirs(PDF_OUTPUT_FOLDER, exist_ok=True)

    pdf_path = os.path.join(UPLOAD_FOLDER, pdf_file.filename)
    pdf_file.save(pdf_path)

    doc = fitz.open(pdf_path)
    sources = []
    diagram_count = 1

    # For each page, render as image, run YOLO, crop and save diagrams
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=300)
        img_bytes = pix.tobytes("png")
        img_np = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), cv2.IMREAD_COLOR)
        h, w, _ = img_np.shape
        results = model(img_np, conf=CONFIDENCE_THRESHOLD)
        for result in results:
            boxes = result.boxes.xyxy.cpu().numpy()
            for box in boxes:
                x1, y1, x2, y2 = map(int, box)
                x1 = max(0, x1 - PADDING)
                y1 = max(0, y1 - PADDING)
                x2 = min(w, x2 + PADDING)
                y2 = min(h, y2 + PADDING)
                cropped_diagram = img_np[y1:y2, x1:x2]
                out_path = os.path.join(PDF_OUTPUT_FOLDER, f"diagram_{diagram_count}.png")
                cv2.imwrite(out_path, cropped_diagram)
                pdf_source = f"{pdf_file.filename}_page_{page_num + 1}_diagram_{diagram_count}"
                sources.append(pdf_source)
                diagram_count += 1
    doc.close()

    image_urls = [url_for('static', filename=f'diagrams_pdf_filtered/{img}') for img in sorted(os.listdir(PDF_OUTPUT_FOLDER))]
    if not image_urls:
        return render_template(
            "message.html",
            status="error",
            message="⚠️ No diagrams were detected in the PDF.",
            redirect_url=url_for("index"),
            redirect_message="Try another PDF"
        )
    return render_template('preview.html', image_urls=image_urls, sources=sources)



# --- API: Download selected images, save to cleaned or extracted folder, and log ---
@app.route('/download_selected', methods=['POST'])
def download_selected():
    selected = request.form.getlist('selected_images')
    if not selected:
        return render_template("message.html", status="error", message="⚠️ No images selected for download.", redirect_url=url_for('index'))

    # If saving to cleaned local folder
    if request.form.get('clean_local'):
        extracted_dir = r'E:\Cleaned data'
        os.makedirs(extracted_dir, exist_ok=True)
        added = False
        for idx, relative_path in enumerate(selected):
            src = os.path.join('static', relative_path)
            original_name = os.path.basename(relative_path)
            dst = os.path.join(extracted_dir, original_name)
            cropped_key = f"cropped_image_data_{idx}"
            cropped_data = request.form.get(cropped_key)
            if cropped_data and cropped_data.startswith("data:image"):
                header, encoded = cropped_data.split(",", 1)
                img_bytes = base64.b64decode(encoded)
                with open(dst, "wb") as f:
                    f.write(img_bytes)
                added = True
                continue
            if os.path.abspath(src) == os.path.abspath(dst):
                continue
            if os.path.exists(src):
                shutil.copyfile(src, dst)
                added = True
        if not added:
            return render_template("message.html", status="error", message="⚠️ No valid images found to copy.", redirect_url=url_for('index'))
        return render_template("message.html", status="success", message="✅ Selected images saved with original names in Cleaned data!", redirect_url=url_for('index'))

    # Otherwise, save to extracted diagrams folder and log to Excel
    extracted_dir = os.path.join('results', 'Diagrams_Extracted')
    os.makedirs(extracted_dir, exist_ok=True)
    existing_numbers = [int(name) for name, ext in (os.path.splitext(f) for f in os.listdir(extracted_dir)) if name.isdigit()]
    next_number = max(existing_numbers, default=19999) + 1
    added = False
    for idx, relative_path in enumerate(selected):
        src = os.path.join('static', relative_path)
        ext = os.path.splitext(relative_path)[-1]
        dst = os.path.join(extracted_dir, f"{next_number}{ext}")
        cropped_key = f"cropped_image_data_{idx}"
        cropped_data = request.form.get(cropped_key)
        if cropped_data and cropped_data.startswith("data:image"):
            header, encoded = cropped_data.split(",", 1)
            img_bytes = base64.b64decode(encoded)
            with open(dst, "wb") as f:
                f.write(img_bytes)
            added = True
            source = request.form.get(f'source_{idx}', '')
            log_image_to_excel(os.path.basename(dst), source)
            next_number += 1
            continue
        if os.path.abspath(src) == os.path.abspath(dst):
            continue
        if os.path.exists(src):
            shutil.copyfile(src, dst)
            added = True
            source = request.form.get(f'source_{idx}', '')
            log_image_to_excel(os.path.basename(dst), source)
            next_number += 1
    if not added:
        return render_template("message.html", status="error", message="⚠️ No valid images found to copy.", redirect_url=url_for('index'))
    return render_template("message.html", status="success", message="✅ Selected images saved in order!", redirect_url=url_for('index'))



# --- API: Upload a folder of images and preview them ---
@app.route('/upload_folder', methods=['POST'])
def upload_folder():
    files = request.files.getlist('images')
    if not files:
        return redirect(url_for('index'))
    FOLDER_UPLOAD_OUTPUT = 'static/diagrams_folder_filtered'
    shutil.rmtree(FOLDER_UPLOAD_OUTPUT, ignore_errors=True)
    os.makedirs(FOLDER_UPLOAD_OUTPUT, exist_ok=True)
    for file in files:
        filename = file.filename
        ext = os.path.splitext(filename)[-1].lower()
        if ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff']:
            out_path = os.path.join(FOLDER_UPLOAD_OUTPUT, filename)
            file.save(out_path)
    image_urls = [url_for('static', filename=f'diagrams_folder_filtered/{img}') for img in sorted(os.listdir(FOLDER_UPLOAD_OUTPUT))]
    return render_template('preview.html', image_urls=image_urls, clean_local=True, sources=[])



# --- API: Save a cleaned image to local folder ---
@app.route('/save_cleaned_image', methods=['POST'])
def save_cleaned_image():
    data = request.json
    filename = data.get('filename')
    cropped_data = data.get('cropped_data')
    if not filename or not cropped_data:
        return {"status": "error", "message": "Missing data"}, 400
    extracted_dir = r'E:\Cleaned data'
    os.makedirs(extracted_dir, exist_ok=True)
    dst = os.path.join(extracted_dir, filename)
    if cropped_data.startswith("data:image"):
        header, encoded = cropped_data.split(",", 1)
        img_bytes = base64.b64decode(encoded)
        with open(dst, "wb") as f:
            f.write(img_bytes)
        return {"status": "success", "message": "Image saved"}
    else:
        return {"status": "error", "message": "Invalid image data"}, 400



# --- API: Delete a cleaned image from local folder ---
@app.route('/delete_cleaned_image', methods=['POST'])
def delete_cleaned_image():
    data = request.json
    filename = data.get('filename')
    if not filename:
        return {"status": "error", "message": "Missing filename"}, 400
    extracted_dir = r'E:\Cleaned data'
    file_path = os.path.join(extracted_dir, filename)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            return {"status": "success", "message": "File deleted"}
        except Exception as e:
            return {"status": "error", "message": str(e)}, 500
    else:
        return {"status": "not_found", "message": "File does not exist"}, 200



# --- Main entry point ---
if __name__ == '__main__':
    app.run(debug=True, port=8080)